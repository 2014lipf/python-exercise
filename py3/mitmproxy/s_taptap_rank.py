
#author lamyoung

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
import random
import time

ICON_TEMP='icon'
if os.path.isdir(ICON_TEMP)==False:
	os.mkdir(ICON_TEMP)


TITLE_LIST=['排名','id','游戏名称','地址','评分','关注','下载量','icon']
wb = Workbook()
dest_filename = 'taptap_rank.xlsx'
ws1 = wb.active
for col in range(0, len(TITLE_LIST)):
    _ = ws1.cell(column=col+1, row=1, value="{0}".format(TITLE_LIST[col]))
ws1.column_dimensions['C'].width=30
ws1.column_dimensions['D'].width=40
row_count = 1

def decode(content):
	global row_count,ws1
	if content['success']:
		data_list=content['data']['list']
		for data in data_list:
			print(f'------->{row_count}')
			row_count = row_count+1
			ws1.row_dimensions[row_count].height=40

			data_id = data['id']
			data_title = data['title']
			data_stat = data['stat']
			link = f'https://www.taptap.com/app/{data_id}/'
			tags = ','.join([tag['value'] for tag in data['tags']])
			icon_url = data['icon']['url']
			icon_path = os.path.join('.',ICON_TEMP, f'{data_id}.png')
			if os.path.isfile(icon_path)==False:
				time.sleep(random.random()*2)
				icon_r = requests.get(icon_url);
				with open(icon_path, 'wb') as fd:
				    fd.write(icon_r.content)
			score = data_stat['rating']['score']
			fans_count = data_stat['fans_count'] #关注
			hits_total = data_stat['hits_total'] #下载 

			_ = ws1.cell(column=1, row=row_count, value="{0}".format(row_count-1))
			_ = ws1.cell(column=2, row=row_count, value="{0}".format(data_id))
			_ = ws1.cell(column=3, row=row_count, value="{0}".format(data_title))
			_ = ws1.cell(column=4, row=row_count, value="{0}".format(link))
			_ = ws1.cell(column=5, row=row_count, value="{0}".format(score))
			_ = ws1.cell(column=6, row=row_count, value="{0}".format(fans_count))
			_ = ws1.cell(column=7, row=row_count, value="{0}".format(hits_total))
			img = Image(icon_path)
			img.width=img.height=50
			ws1.add_image(img, f'H{row_count}')
			

headers = {
	'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.87 Safari/537.36',
	'Accept': 'application/json' 
}

def scrapy(index):
	params = {
		'X-UA': 'V=1&PN=TapTap&VN_CODE=536&LOC=CN&LANG=zh_CN&CH=tencent',#'V%3D1%26PN%3DTapTap%26VN_CODE%3D536%26LOC%3DCN%26LANG%3Dzh_CN%26CH%3Dtencent',
		'tag': '单机',#'%E5%8D%95%E6%9C%BA',
		'sort': 'hits',
		'from': f'{index*10}',
		'limit': '10',
	}
	requests_url = 'https://api.taptapdada.com/app-tag/v1/by-tag'
	requests_page=requests.get(requests_url,headers=headers,params=params);
	if requests_page.status_code == 200:
		decode(requests_page.json())
	else:
		print(f'{requests_url}-{requests_page.status_code}')

for x in range(0,10):
	scrapy(x)

wb.save(filename = dest_filename)
